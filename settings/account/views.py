import openpyxl
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import CreateView, UpdateView, DeleteView, ListView
from django.urls import reverse_lazy
from tournament.models import Tournament, Region
from tournament.forms import TournamentForm
from players.models import ChessPlayers
from utils import mongo_db, set_settings, delete_setting


def add_tournament(request):
    if request.method == "POST":
        form = TournamentForm(request.POST)
        if form.is_valid():
            tournament = form.save(commit=False)  # Создание объекта турнира без сохранения в базу данных
            tournament.organizer = request.user  # Назначение текущего пользователя как организатора
            tournament.save()  # Сохранение объекта турнира в базу данных
            return HttpResponseRedirect(reverse_lazy('database'))

    else:
        form = TournamentForm()
    return render(request, 'account/create_tournaments.html', {
        'form': form,
    })


class TournamentListView(ListView):
    model = Tournament
    template_name = 'account/tournaments.html'
    context_object_name = 'tournaments'

    extra_context = {'regions': Region.objects.all()}


class TournamentUpdateView(UpdateView):
    model = Tournament
    template_name = 'account/tournaments.html'
    fields = ['title', 'description', 'start_date', 'end_date', 'gender', 'min_age', 'max_age', 'region']
    success_url = reverse_lazy('database')
    context_object_name = 'tournament'


class TournamentDeleteView(DeleteView):
    model = Tournament
    template_name = 'account/tournaments.html'
    success_url = reverse_lazy('account:tournaments')


def update_tournament_players(request):
    if request.method == 'POST':
        tournament_id = request.POST.get('tournament_id')
        selected_players = request.POST.getlist('players')
        tournament = get_object_or_404(Tournament, id=tournament_id)
        tournament.players.set(selected_players)
        return redirect('account:tournaments')


def get_tournament_players(request, tournament_id):
    tournament = get_object_or_404(Tournament, id=tournament_id)
    players = tournament.players.all()
    players_data = [{'id': player.id, 'first_name': player.first_name, 'last_name': player.last_name} for player in
                    players]
    return JsonResponse({'players': players_data})


def remove_player_from_tournament(request, tournament_id, player_id):
    tournament = get_object_or_404(Tournament, id=tournament_id)
    player = get_object_or_404(ChessPlayers, id=player_id)

    if request.method == 'POST':
        # Удаление игрока из турнира
        tournament.players.remove(player)
        return redirect('account:tournaments')


def download_players_excel(request, tournament_id):
    tournament = Tournament.objects.get(id=tournament_id)
    players = tournament.players.all()

    # Создаем новый workbook и worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Players"

    # Записываем заголовки столбцов
    ws['A1'] = "Имя"
    ws['B1'] = "Фамилия"

    # Записываем данные игроков
    row = 2
    for player in players:
        ws[f'A{row}'] = player.first_name
        ws[f'B{row}'] = player.last_name
        row += 1

    # Создаем HttpResponse с данными в формате Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=players_{tournament.title}.xlsx'

    # Сохраняем workbook в HttpResponse
    wb.save(response)

    return response


def notifications(request):
    if request.method == 'POST':
        unsubscribe = request.POST.get('unsubscribe')
        email = request.POST.get('email')
        telegram = request.POST.get('telegram')
        notification_type = request.POST.get('notification_type')
        data = {'user_id': request.user.id}
        if unsubscribe == 'on':
            delete_setting(request.user.id)
        else:
            if email:
                data['email'] = email
            if telegram:
                data['telegram'] = telegram
            if notification_type:
                data['notification_type'] = notification_type
            if data:
                set_settings(request.user.id, data)
    context = {}
    return render(request, 'account/check.html', context=context)
